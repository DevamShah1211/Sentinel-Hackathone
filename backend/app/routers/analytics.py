"""
Analytics router — dashboard summaries + output report generation (XLSX + PDF).
The output report is a required submission artefact per the playbook.
"""
import io
import logging
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select, desc, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models import Alert, Camera, Detection, WatchlistEntry

logger = logging.getLogger("sentinel.analytics")
router = APIRouter()


@router.get("/summary", summary="Platform-wide dashboard summary")
async def summary(db: AsyncSession = Depends(get_db)):
    # Camera stats
    cam_total = await db.scalar(select(func.count(Camera.id)))
    cam_live  = await db.scalar(select(func.count(Camera.id)).where(Camera.is_live == True))

    # Detection stats (last 24h)
    since = datetime.now(timezone.utc) - timedelta(hours=24)
    det_24h = await db.scalar(
        select(func.count(Detection.id)).where(Detection.detected_at >= since)
    )
    det_total = await db.scalar(select(func.count(Detection.id)))

    # Alert stats
    alerts_new  = await db.scalar(select(func.count(Alert.id)).where(Alert.status == "new"))
    alerts_total = await db.scalar(select(func.count(Alert.id)))

    # Watchlist
    wl_active = await db.scalar(
        select(func.count(WatchlistEntry.id)).where(WatchlistEntry.active == True)
    )

    # Unique plates detected (last 24h)
    unique_plates = await db.scalar(
        select(func.count(func.distinct(Detection.plate_text)))
        .where(Detection.detected_at >= since)
    )

    return {
        "cameras": {"total": cam_total, "live": cam_live, "offline": cam_total - cam_live},
        "detections": {"total": det_total, "last_24h": det_24h, "unique_plates_24h": unique_plates},
        "alerts": {"total": alerts_total, "new": alerts_new},
        "watchlist": {"active_entries": wl_active},
        "generated_at": datetime.now(timezone.utc).isoformat(),
    }


@router.get("/top-plates", summary="Most frequently detected plates")
async def top_plates(db: AsyncSession = Depends(get_db), limit: int = 20):
    result = await db.execute(
        select(Detection.plate_text, func.count(Detection.id).label("count"))
        .group_by(Detection.plate_text)
        .order_by(desc(func.count(Detection.id)))
        .limit(limit)
    )
    return [{"plate_text": r[0], "count": r[1]} for r in result.all()]


@router.get("/detections-by-hour", summary="Detections per camera per hour (last 24h)")
async def detections_by_hour(db: AsyncSession = Depends(get_db)):
    since = datetime.now(timezone.utc) - timedelta(hours=24)
    result = await db.execute(
        select(
            func.date_trunc("hour", Detection.detected_at).label("hour"),
            func.count(Detection.id).label("count"),
        )
        .where(Detection.detected_at >= since)
        .group_by(text("1"))
        .order_by(text("1"))
    )
    return [{"hour": str(r[0]), "count": r[1]} for r in result.all()]


@router.get("/report/xlsx", summary="Download output report as XLSX (required artefact)")
async def download_xlsx(
    db: AsyncSession = Depends(get_db),
    from_dt: Optional[datetime] = Query(None),
    to_dt: Optional[datetime] = Query(None),
    limit: int = Query(5000, le=50000),
):
    """
    Generates the government-feed output report: plate, confidence, camera,
    department, location, UTC timestamp, evidence reference.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    q = (
        select(Detection, Camera.name, Camera.department, Camera.lat, Camera.lon, Camera.address)
        .join(Camera, Detection.camera_id == Camera.id)
        .order_by(desc(Detection.detected_at))
        .limit(limit)
    )
    if from_dt:
        q = q.where(Detection.detected_at >= from_dt)
    if to_dt:
        q = q.where(Detection.detected_at <= to_dt)
    result = await db.execute(q)
    rows = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sentinel ANPR Report"

    # Header row
    headers = [
        "Plate Text", "Confidence %", "Camera Name", "Department",
        "Latitude", "Longitude", "Address", "Detected At (UTC)",
        "Track ID", "Vehicle Type", "Evidence Crop",
    ]
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    hdr_font = Font(bold=True, color="E0E0E0")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    fill_alt = PatternFill("solid", fgColor="F5F5F5")
    for i, row in enumerate(rows, 2):
        det = row[0]
        ws.cell(i, 1, det.plate_text)
        ws.cell(i, 2, round(det.confidence * 100, 1))
        ws.cell(i, 3, row[1])
        ws.cell(i, 4, row[2])
        ws.cell(i, 5, row[3])
        ws.cell(i, 6, row[4])
        ws.cell(i, 7, row[5])
        ws.cell(i, 8, det.detected_at.strftime("%Y-%m-%d %H:%M:%S UTC") if det.detected_at else "")
        ws.cell(i, 9, det.track_id)
        ws.cell(i, 10, det.vehicle_type)
        ws.cell(i, 11, det.crop_uri)
        if i % 2 == 0:
            for c in range(1, 12):
                ws.cell(i, c).fill = fill_alt

    # Column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"sentinel_report_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

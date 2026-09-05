"""
Output report generation — the required submission artefact.

The government-feed demonstration must be submitted "along with an output report
showing detected vehicles or number plates with corresponding timestamps". This
module builds that report from the database in both XLSX and PDF, so it always
reflects what the system actually indexed rather than anything assembled by hand.

Both formats carry the same content: a provenance header saying where the data
came from and how locations were derived, a summary, and the detection table with
plate, confidence, camera, department, coordinates and UTC timestamp.
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Sequence

from app.gap_analysis import NOMINAL_COVERAGE_KM

BRAND_DARK = "1A2332"
BRAND_ACCENT = "2F6FEB"


@dataclass
class ReportRow:
    """One detection, flattened with its camera context."""
    plate_text: str
    confidence: float
    camera_native_id: str
    camera_name: str
    department: str
    lat: float | None
    lon: float | None
    address: str | None
    detected_at: datetime | None
    pts_ms: int | None
    track_id: str | None
    crop_uri: str | None
    geo_source: str | None = None

    @property
    def timestamp_utc(self) -> str:
        if not self.detected_at:
            return ""
        return self.detected_at.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    @property
    def coordinates(self) -> str:
        if self.lat is None or self.lon is None:
            return "not resolved"
        return f"{self.lat:.5f}, {self.lon:.5f}"


@dataclass
class ReportMeta:
    """Context printed at the top of the report so the numbers can be interpreted."""
    title: str = "Sentinel — ANPR Output Report"
    subtitle: str = "Gujarat CCTV Integration Hackathon 2026 · Model 1 + Model 2"
    source: str = "Sentinel sandbox grid (cctv.corp8.cloud)"
    generated_at: datetime | None = None
    total_detections: int = 0
    unique_plates: int = 0
    cameras_covered: int = 0
    watchlist_alerts: int = 0
    window_from: datetime | None = None
    window_to: datetime | None = None
    notes: Sequence[str] = ()

    def __post_init__(self) -> None:
        if self.generated_at is None:
            self.generated_at = datetime.now(timezone.utc)

    @property
    def window(self) -> str:
        if self.window_from and self.window_to:
            return (f"{self.window_from:%Y-%m-%d %H:%M} to "
                    f"{self.window_to:%Y-%m-%d %H:%M} UTC")
        return "all indexed detections"


DEFAULT_NOTES = (
    "Plates are read by pretrained open-source models (YOLOv9 plate detector, "
    "cct-s-v2 OCR) running locally on CPU. No video frame leaves the deployment "
    "and no external API is used in the detection or alerting path.",
    "Each row is one vehicle pass, not one frame: reads are aggregated across a "
    "vehicle's track by confidence-weighted per-character voting, then validated "
    "against Indian plate grammar.",
    "The sandbox catalogue publishes only camera id and name. Coordinates are "
    "resolved by hand-verification or OpenStreetMap geocoding and each camera "
    "records which method was used; they are approximate site locations, not a "
    "surveyed asset register.",
)


# ─── XLSX ─────────────────────────────────────────────────────────────────────

def build_xlsx(rows: Sequence[ReportRow], meta: ReportMeta) -> bytes:
    """Render the report as a formatted workbook."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = openpyxl.Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────────
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = meta.title
    summary["A1"].font = Font(size=16, bold=True, color=BRAND_DARK)
    summary["A2"] = meta.subtitle
    summary["A2"].font = Font(size=11, color="555555")

    facts = [
        ("Generated (UTC)", meta.generated_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Data source", meta.source),
        ("Reporting window", meta.window),
        ("Total detections", meta.total_detections),
        ("Unique plates", meta.unique_plates),
        ("Cameras covered", meta.cameras_covered),
        ("Watchlist alerts", meta.watchlist_alerts),
    ]
    for offset, (label, value) in enumerate(facts, start=4):
        summary.cell(offset, 1, label).font = Font(bold=True)
        summary.cell(offset, 2, value)

    row_cursor = 4 + len(facts) + 1
    summary.cell(row_cursor, 1, "Method and limitations").font = Font(bold=True, size=12)
    for offset, note in enumerate(meta.notes or DEFAULT_NOTES, start=row_cursor + 1):
        cell = summary.cell(offset, 1, f"• {note}")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        summary.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=6)
        summary.row_dimensions[offset].height = 30
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 60

    # ── Detections sheet ─────────────────────────────────────────────────────
    sheet = workbook.create_sheet("Detections")
    headers = [
        "Plate", "Confidence %", "Camera ID", "Camera Name", "Department",
        "Latitude", "Longitude", "Location", "Detected At (UTC)",
        "Stream PTS (ms)", "Track ID", "Evidence Crop", "Location Source",
    ]
    header_fill = PatternFill("solid", fgColor=BRAND_DARK)
    header_font = Font(bold=True, color="FFFFFF")
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(1, column, title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 24
    sheet.freeze_panes = "A2"

    stripe = PatternFill("solid", fgColor="F2F5FA")
    for index, row in enumerate(rows, start=2):
        values = [
            row.plate_text,
            round(row.confidence * 100, 1),
            row.camera_native_id,
            row.camera_name,
            row.department,
            row.lat,
            row.lon,
            row.address,
            row.timestamp_utc,
            row.pts_ms,
            row.track_id,
            row.crop_uri,
            row.geo_source,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(index, column, value)
            if index % 2 == 0:
                cell.fill = stripe
        sheet.cell(index, 1).font = Font(bold=True)

    widths = [14, 13, 11, 30, 22, 11, 11, 40, 22, 15, 20, 34, 15]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 2)}"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ─── PDF ──────────────────────────────────────────────────────────────────────

def build_pdf(rows: Sequence[ReportRow], meta: ReportMeta,
              max_rows: int = 600) -> bytes:
    """
    Render the report as a paginated PDF.

    Long runs are truncated with an explicit note rather than silently cut — the
    XLSX carries the complete set.
    """
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=14 * mm,
        title=meta.title, author="Sentinel CCTV Platform",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("SentinelTitle", parent=styles["Title"],
                                 fontSize=19, textColor=colors.HexColor(f"#{BRAND_DARK}"),
                                 alignment=TA_LEFT, spaceAfter=2)
    subtitle_style = ParagraphStyle("SentinelSubtitle", parent=styles["Normal"],
                                    fontSize=10, textColor=colors.HexColor("#5A6472"),
                                    spaceAfter=10)
    note_style = ParagraphStyle("SentinelNote", parent=styles["Normal"],
                                fontSize=8.5, leading=12,
                                textColor=colors.HexColor("#333A44"))
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7.4, leading=9)

    story: list[Any] = [
        Paragraph(meta.title, title_style),
        Paragraph(meta.subtitle, subtitle_style),
    ]

    facts = [
        ["Generated (UTC)", meta.generated_at.strftime("%Y-%m-%d %H:%M:%S"),
         "Total detections", str(meta.total_detections)],
        ["Data source", meta.source, "Unique plates", str(meta.unique_plates)],
        ["Reporting window", meta.window, "Cameras covered", str(meta.cameras_covered)],
        ["", "", "Watchlist alerts", str(meta.watchlist_alerts)],
    ]
    fact_table = Table(facts, colWidths=[34 * mm, 108 * mm, 36 * mm, 34 * mm])
    fact_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#2A3340")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("LINEBELOW", (0, -1), (-1, -1), 0.6, colors.HexColor("#C9D2DE")),
    ]))
    story += [fact_table, Spacer(1, 7)]

    story.append(Paragraph("<b>Method and limitations</b>", note_style))
    for note in (meta.notes or DEFAULT_NOTES):
        story.append(Paragraph(f"• {note}", note_style))
    story.append(Spacer(1, 9))

    headers = ["Plate", "Conf %", "Cam", "Camera / Location", "Department",
               "Coordinates", "Detected At (UTC)", "Track"]
    data: list[list[Any]] = [headers]
    truncated = list(rows)[:max_rows]
    for row in truncated:
        location = row.camera_name
        if row.address:
            location = f"{row.camera_name}<br/><font color='#6B7480'>{row.address}</font>"
        data.append([
            row.plate_text,
            f"{row.confidence * 100:.0f}",
            row.camera_native_id,
            Paragraph(location, cell_style),
            Paragraph(row.department or "", cell_style),
            row.coordinates,
            row.timestamp_utc,
            Paragraph(row.track_id or "", cell_style),
        ])

    table = Table(
        data, repeatRows=1,
        colWidths=[24 * mm, 13 * mm, 13 * mm, 72 * mm, 32 * mm, 34 * mm, 38 * mm, 26 * mm],
    )
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{BRAND_DARK}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.4),
        ("ALIGN", (1, 1), (2, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F2F5FA")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D5DCE6")),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]))
    story.append(table)

    if len(rows) > max_rows:
        story += [
            Spacer(1, 6),
            Paragraph(
                f"Showing the {max_rows} most recent of {len(rows)} detections. "
                "The XLSX edition of this report contains the complete set.",
                note_style,
            ),
        ]

    def decorate(canvas, doc) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(colors.HexColor("#7A828E"))
        canvas.drawString(12 * mm, 8 * mm,
                          "Sentinel CCTV Platform · generated from the detection index")
        canvas.drawRightString(landscape(A4)[0] - 12 * mm, 8 * mm, f"Page {doc.page}")
        canvas.restoreState()

    document.build(story, onFirstPage=decorate, onLaterPages=decorate)
    return buffer.getvalue()


# ─── Gap-analysis report (Model 1 deliverable) ────────────────────────────────

SEVERITY_FILL = {
    "critical": "C0392B",
    "high": "E67E22",
    "medium": "F1C40F",
    "warning": "E67E22",
    "info": "5D6D7E",
    "covered": "27AE60",
}


def build_gap_xlsx(report) -> bytes:
    """
    Render the coverage and condition analysis as a workbook.

    Three sheets, because they answer three different questions: what the estate
    looks like overall, which districts are thin, and which individual cameras
    need attention.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor=BRAND_DARK)
    header_font = Font(bold=True, color="FFFFFF")

    def write_header(sheet, titles: list[str]) -> None:
        for column, title in enumerate(titles, start=1):
            cell = sheet.cell(1, column, title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 22
        sheet.freeze_panes = "A2"

    # ── Summary ──────────────────────────────────────────────────────────────
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "Sentinel — Coverage Gap Analysis"
    summary["A1"].font = Font(size=16, bold=True, color=BRAND_DARK)
    summary["A2"] = "Gujarat CCTV Integration Hackathon 2026 · Model 1 deliverable"
    summary["A2"].font = Font(size=11, color="555555")

    facts = [
        ("Generated (UTC)", report.generated_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Cameras in registry", report.total_cameras),
        ("Cameras with resolved locations", report.located_cameras),
        ("Districts assessed", report.districts_total),
        ("Districts with coverage", report.districts_covered),
        ("Districts without coverage", report.districts_uncovered),
        ("Coverage", f"{report.coverage_percent}%"),
        ("Ageing / condition findings", len(report.ageing)),
    ]
    for offset, (label, value) in enumerate(facts, start=4):
        summary.cell(offset, 1, label).font = Font(bold=True)
        summary.cell(offset, 2, value)

    notes_row = 4 + len(facts) + 1
    summary.cell(notes_row, 1, "Method and limitations").font = Font(bold=True, size=12)
    notes = [
        "A district counts as covered when a registered camera lies within "
        f"{NOMINAL_COVERAGE_KM:.0f} km of its centre. This is a spatial test, not a "
        "count: several cameras on one junction do not cover a district.",
        "Distances are great-circle from the district centre to the nearest camera, "
        "so they understate road distance and the real gap is usually larger.",
        "Condition findings use only what the registry holds. Where a field is "
        "absent the camera is reported as unknown, never assumed healthy.",
        "The sandbox catalogue publishes 30 cameras for demonstration. Uncovered "
        "districts here reflect that sample, not the true Gujarat estate.",
    ]
    for offset, note in enumerate(notes, start=notes_row + 1):
        cell = summary.cell(offset, 1, f"• {note}")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        summary.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=6)
        summary.row_dimensions[offset].height = 30
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 58

    # ── Coverage by district ─────────────────────────────────────────────────
    sheet = workbook.create_sheet("Coverage by district")
    write_header(sheet, ["District", "Severity", "Cameras within "
                         f"{NOMINAL_COVERAGE_KM:.0f} km", "Nearest camera (km)",
                         "Nearest camera", "Latitude", "Longitude"])
    for index, entry in enumerate(
            sorted(report.coverage, key=lambda c: -(c.nearest_camera_km or 0)), start=2):
        sheet.cell(index, 1, entry.district)
        severity_cell = sheet.cell(index, 2, entry.severity.title())
        severity_cell.fill = PatternFill(
            "solid", fgColor=SEVERITY_FILL.get(entry.severity, "5D6D7E"))
        severity_cell.font = Font(bold=True, color="FFFFFF")
        sheet.cell(index, 3, entry.cameras_within_radius)
        sheet.cell(index, 4, entry.nearest_camera_km)
        sheet.cell(index, 5, entry.nearest_camera_name)
        sheet.cell(index, 6, entry.lat)
        sheet.cell(index, 7, entry.lon)
    for column, width in enumerate([22, 12, 20, 20, 34, 12, 12], start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    # ── Ageing infrastructure ────────────────────────────────────────────────
    sheet = workbook.create_sheet("Infrastructure condition")
    write_header(sheet, ["Camera ID", "Name", "District", "Severity",
                         "Issue", "Evidence"])
    for index, finding in enumerate(report.ageing, start=2):
        sheet.cell(index, 1, finding.native_id)
        sheet.cell(index, 2, finding.name)
        sheet.cell(index, 3, finding.district)
        severity_cell = sheet.cell(index, 4, finding.severity.title())
        severity_cell.fill = PatternFill(
            "solid", fgColor=SEVERITY_FILL.get(finding.severity, "5D6D7E"))
        severity_cell.font = Font(bold=True, color="FFFFFF")
        sheet.cell(index, 5, finding.issue)
        sheet.cell(index, 6, finding.detail).alignment = Alignment(wrap_text=True)
    for column, width in enumerate([14, 30, 18, 12, 36, 60], start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
